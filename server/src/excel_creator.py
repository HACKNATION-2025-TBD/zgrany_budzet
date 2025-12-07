import openpyxl
from dataclasses import dataclass, field
from typing import List, Optional
import os


# --- 1. TWOJE NOWE DATACLASSES ---

@dataclass
class DaneFinansoweRoku:
    """
    Reprezentuje kolumny finansowe powtarzalne dla każdego roku (n, n+1...).
    """
    potrzeby: float = 0.0  # Potrzeby finansowe
    limit: float = 0.0  # Limit wydatków
    niezabezpieczone: float = 0.0  # Kwota niezabezpieczona w limicie

    @property
    def bilans(self) -> float:
        """Zwraca różnicę: Limit - Potrzeby (pomocnicze)"""
        return self.limit - self.potrzeby


@dataclass
class BudgetEntry:
    # --- 1. Klasyfikacja Budżetowa (Stała dla wiersza) ---
    czesc_budzetowa: str = ""
    dzial: str = ""
    rozdzial: str = ""
    paragraf: str = ""
    zrodlo_finansowania: str = ""
    grupa_wydatkow: str = ""

    # --- 2. Opis Zadania i Organizacja (Stałe dla wiersza) ---
    bz_pelna_szczegolowosc: str = ""
    bz_kody: str = ""
    nazwa_programu_projektu: str = ""
    nazwa_komorki_org: str = ""
    plan_wi: str = ""
    dysponent_srodkow: str = ""
    budzet: str = ""  # np. "Centralny"

    nazwa_zadania: str = ""
    szczegolowe_uzasadnienie: str = ""
    obszar_dzialalnosci: str = ""  # np. cyberbezpieczeństwo

    # --- 3. FINANSE (Rozbite na lata n, n+1, n+2, n+3) ---
    # Używamy default_factory, aby każdy rok był niezależnym obiektem
    n: DaneFinansoweRoku = field(default_factory=DaneFinansoweRoku)  # Rok bieżący (np. 2026)
    n1: DaneFinansoweRoku = field(default_factory=DaneFinansoweRoku)  # Rok n+1 (np. 2027)
    n2: DaneFinansoweRoku = field(default_factory=DaneFinansoweRoku)  # Rok n+2 (np. 2028)
    n3: DaneFinansoweRoku = field(default_factory=DaneFinansoweRoku)  # Rok n+3 (np. 2029)

    # --- 4. Dane o Umowach (Zazwyczaj dotyczą całości zadania) ---
    kwota_umowy_wniosku: float = 0.0
    nr_umowy_wniosku: Optional[str] = None
    dotacja_kontrahent: Optional[str] = None
    podstawa_prawna_dotacji: Optional[str] = None
    uwagi: Optional[str] = None


# --- 2. FUNKCJA EKSPORTUJĄCA DO EXCELA ---

def export_entries_to_excel(template_path: str, output_path: str, entries: List[BudgetEntry], start_row: int = 2):
    """
    Wypełnia szablon Excela danymi z listy obiektów BudgetEntry, obsługując rozbudowaną strukturę.
    """

    # Lista nagłówków (używana tylko jeśli tworzymy nowy plik)
    headers = [
        # Sekcja 1
        "Część", "Dział", "Rozdział", "Paragraf", "Źródło fin.", "Grupa wydatków",
        # Sekcja 2
        "BZ (Pełny)", "BZ (Kody)", "Program/Projekt", "Komórka Org.", "Plan WI", "Dysponent", "Budżet",
        # Sekcja 3 (Zadanie)
        "Nazwa Zadania", "Uzasadnienie", "Obszar działania",
        # Sekcja 4 (Finanse - Rok N)
        "Potrzeby (N)", "Limit (N)", "Niezabezpieczone (N)",
        # Finanse - Rok N+1
        "Potrzeby (N+1)", "Limit (N+1)", "Niezabezpieczone (N+1)",
        # Finanse - Rok N+2
        "Potrzeby (N+2)", "Limit (N+2)", "Niezabezpieczone (N+2)",
        # Finanse - Rok N+3
        "Potrzeby (N+3)", "Limit (N+3)", "Niezabezpieczone (N+3)",
        # Sekcja 5 (Umowy/Inne)
        "Kwota Umowy/Wniosku", "Nr Umowy", "Kontrahent (Dotacja)", "Podstawa Prawna", "Uwagi"
    ]

    if not os.path.exists(template_path):
        print(f"⚠️ Plik {template_path} nie istnieje. Tworzę nowy arkusz z nagłówkami.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        start_row = 2  # Dane zaczynamy od 2 wiersza
    else:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

    print(f"Zapisywanie {len(entries)} wierszy do {output_path}...")

    current_row = start_row

    for entry in entries:
        # --- SEKCJA 1: KLASYFIKACJA ---
        ws.cell(row=current_row, column=1, value=entry.czesc_budzetowa)
        ws.cell(row=current_row, column=2, value=entry.dzial)
        ws.cell(row=current_row, column=3, value=entry.rozdzial)
        ws.cell(row=current_row, column=4, value=entry.paragraf)
        ws.cell(row=current_row, column=5, value=entry.zrodlo_finansowania)
        ws.cell(row=current_row, column=6, value=entry.grupa_wydatkow)

        # --- SEKCJA 2: ORGANIZACJA I BZ ---
        ws.cell(row=current_row, column=7, value=entry.bz_pelna_szczegolowosc)
        ws.cell(row=current_row, column=8, value=entry.bz_kody)
        ws.cell(row=current_row, column=9, value=entry.nazwa_programu_projektu)
        ws.cell(row=current_row, column=10, value=entry.nazwa_komorki_org)
        ws.cell(row=current_row, column=11, value=entry.plan_wi)
        ws.cell(row=current_row, column=12, value=entry.dysponent_srodkow)
        ws.cell(row=current_row, column=13, value=entry.budzet)

        # --- SEKCJA 3: ZADANIE ---
        ws.cell(row=current_row, column=14, value=entry.nazwa_zadania)
        ws.cell(row=current_row, column=15, value=entry.szczegolowe_uzasadnienie)
        ws.cell(row=current_row, column=16, value=entry.obszar_dzialalnosci)

        # --- SEKCJA 4: FINANSE (Lata N do N+3) ---

        # Rok N (2026)
        ws.cell(row=current_row, column=17, value=entry.n.potrzeby)
        ws.cell(row=current_row, column=18, value=entry.n.limit)
        ws.cell(row=current_row, column=19, value=entry.n.niezabezpieczone)

        # Rok N+1 (2027)
        ws.cell(row=current_row, column=20, value=entry.n1.potrzeby)
        ws.cell(row=current_row, column=21, value=entry.n1.limit)
        ws.cell(row=current_row, column=22, value=entry.n1.niezabezpieczone)

        # Rok N+2 (2028)
        ws.cell(row=current_row, column=23, value=entry.n2.potrzeby)
        ws.cell(row=current_row, column=24, value=entry.n2.limit)
        ws.cell(row=current_row, column=25, value=entry.n2.niezabezpieczone)

        # Rok N+3 (2029)
        ws.cell(row=current_row, column=26, value=entry.n3.potrzeby)
        ws.cell(row=current_row, column=27, value=entry.n3.limit)
        ws.cell(row=current_row, column=28, value=entry.n3.niezabezpieczone)

        # --- SEKCJA 5: UMOWY I INNE ---
        ws.cell(row=current_row, column=29, value=entry.kwota_umowy_wniosku)
        ws.cell(row=current_row, column=30, value=entry.nr_umowy_wniosku)
        ws.cell(row=current_row, column=31, value=entry.dotacja_kontrahent)
        ws.cell(row=current_row, column=32, value=entry.podstawa_prawna_dotacji)
        ws.cell(row=current_row, column=33, value=entry.uwagi)

        current_row += 1

    wb.save(output_path)
    print(f"✅ Gotowe! Plik zapisany jako: {output_path}")


# --- PRZYKŁAD UŻYCIA ---

if __name__ == "__main__":
    # 1. Tworzenie przykładowych danych
    wpis_1 = BudgetEntry(
        czesc_budzetowa="27",
        dzial="750",
        rozdzial="75001",
        paragraf="4300",
        grupa_wydatkow="Bieżące",
        nazwa_zadania="Zakup licencji oprogramowania biurowego",
        obszar_dzialalnosci="Koszt funkcjonowania",
        dysponent_srodkow="Departament IT",

        # Finanse na rok N (2026)
        n=DaneFinansoweRoku(potrzeby=120000.0, limit=100000.0, niezabezpieczone=20000.0),

        # Finanse na rok N+1 (2027)
        n1=DaneFinansoweRoku(potrzeby=50000.0, limit=50000.0),

        uwagi="Wymaga pilnego przetargu"
    )

    wpis_2 = BudgetEntry(
        czesc_budzetowa="27",
        dzial="750",
        rozdzial="75001",
        grupa_wydatkow="Majątkowe",
        nazwa_zadania="Budowa serwerowni zapasowej",
        obszar_dzialalnosci="Cyberbezpieczeństwo",

        n=DaneFinansoweRoku(potrzeby=5000000.0, limit=2000000.0, niezabezpieczone=3000000.0),
        n1=DaneFinansoweRoku(potrzeby=2000000.0, limit=2000000.0),
        n2=DaneFinansoweRoku(potrzeby=1000000.0, limit=1000000.0)
    )

    lista_danych = [wpis_1, wpis_2]

    # 2. Definicja plików
    # Jeśli plik "Szablon.xlsx" nie istnieje, skrypt sam go stworzy z nagłówkami
    plik_szablonu = "template.xlsx"
    plik_wynikowy = "Wypelniony_Budzet_Pelny.xlsx"

    # 3. Uruchomienie eksportu
    export_entries_to_excel(plik_szablonu, plik_wynikowy, lista_danych)